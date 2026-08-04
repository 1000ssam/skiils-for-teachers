#!/usr/bin/env python3
"""검수 엑셀에 적으신 판정을 되먹인다 — 검수의 단일 창구.

  out/검수.xlsx 의 '다시 / 확정 / 보류' 칸과 '왜 다시 쓰나요' 칸을 읽어
  out/marks/<학번>.json 에 반영한다. 그다음 render_now.py 를 부르면
  '다시'로 표시하신 사람만 🔁 로 잡힌다.

  다시 → 물린 본문과 사유를 반려 이력에 쌓고 재료 시각을 새로 찍는다(=다시 써야 할 대상이 됨)
  확정 → 완료. 다시는 안 쓴다.
  보류 → 판정을 미룬 것. 다시 쓰지 않고 그대로 둔다.
  빈칸 → 아직 안 정하신 것. **아무것도 건드리지 않는다.**

쓰는 법:
  python3 tool/ingest_review_xlsx.py           # 미리보기(파일을 고치지 않는다)
  python3 tool/ingest_review_xlsx.py --write   # 실제로 반영
"""
import argparse, json, os, sys
from datetime import datetime, timezone

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MARKS = os.path.join(ROOT, "out", "marks")
DRAFTS = os.path.join(ROOT, "out", "drafts_v3b")

COL_HAKBUN, COL_NAME, COL_JUDGE, COL_REASON, COL_SETUK = 1, 2, 3, 4, 5
VALID = {"다시", "확정", "보류"}


def now():
    return datetime.now(timezone.utc).isoformat(timespec="milliseconds").replace("+00:00", "Z")


def draft_setuk(hb):
    p = os.path.join(DRAFTS, f"{hb}.json")
    if not os.path.exists(p):
        return ""
    try:
        return (json.load(open(p, encoding="utf-8")).get("setuk") or "").strip()
    except Exception:
        return ""


def already_ingested(mark, setuk, reason):
    """같은 본문·같은 사유로 이미 물린 적이 있으면 또 쌓지 않는다.

    엑셀을 두 번 되먹이는 건 흔한 일이고, 그때마다 반려가 쌓이면 '3회 이상 물렸다'
    경고가 거짓으로 켜져 다음 초안이 까닭 없이 짧아진다.
    """
    for r in (mark.get("rejects") or []):
        if (r.get("setuk") or "").strip() == setuk and (r.get("reason") or "").strip() == reason:
            return True
    return False


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=os.path.join(ROOT, "out", "검수.xlsx"))
    ap.add_argument("--write", action="store_true", help="실제로 반영한다(없으면 미리보기)")
    a = ap.parse_args()

    if not os.path.exists(a.xlsx):
        sys.exit(f"❌ 검수 엑셀이 없습니다: {a.xlsx}\n   먼저 python3 tool/make_review_xlsx.py 를 돌리세요.")

    ws = openpyxl.load_workbook(a.xlsx, data_only=True).active
    head = [c.value for c in ws[1]]
    if head[:5] != ["학번", "이름", "다시", "왜 다시 쓰나요", "세특"]:
        sys.exit(f"❌ 이 엑셀은 판정 칸이 없는 옛 양식입니다(첫 줄: {head[:5]}).\n"
                 f"   python3 tool/make_review_xlsx.py 로 다시 만든 뒤 적어 주세요.")

    plan, bad, blank = [], [], 0
    for r in range(2, ws.max_row + 1):
        hb = str(ws.cell(r, COL_HAKBUN).value or "").strip()
        if not hb:
            continue
        judge = str(ws.cell(r, COL_JUDGE).value or "").strip()
        reason = str(ws.cell(r, COL_REASON).value or "").strip()
        name = str(ws.cell(r, COL_NAME).value or "").strip()
        if not judge:
            blank += 1
            if reason:      # 사유만 적고 판정을 안 고르신 경우 — 조용히 버리면 안 된다
                bad.append((hb, name, "사유는 적으셨는데 '다시/확정/보류'를 안 고르셨습니다"))
            continue
        if judge not in VALID:
            bad.append((hb, name, f"'{judge}' 는 못 알아듣습니다. 다시/확정/보류 중에서 골라 주세요"))
            continue
        plan.append((hb, name, judge, reason))

    print(f"검수 엑셀: {a.xlsx}")
    print(f"  판정하신 것 {len(plan)}명 · 비워 두신 것 {blank}명")
    for hb, name, why in bad:
        print(f"  ⚠️ {hb} {name} — {why}")

    if bad:
        print("\n위 줄을 고쳐 주세요. 그대로 두면 그 학생은 건너뜁니다.")

    counts, changes = {}, []
    for hb, name, judge, reason in plan:
        mp = os.path.join(MARKS, f"{hb}.json")
        if not os.path.exists(mp):
            print(f"  ⚠️ {hb} {name} — 표시 기록이 없어 건너뜁니다")
            continue
        m = json.load(open(mp, encoding="utf-8"))
        t = now()

        if judge == "확정":
            if m.get("approved"):
                continue
            m["approved"], m["approved_at"], m["held"] = True, t, False
            changes.append((mp, m)); counts["확정"] = counts.get("확정", 0) + 1

        elif judge == "보류":
            if m.get("held"):
                continue
            m["held"], m["held_at"], m["approved"] = True, t, False
            changes.append((mp, m)); counts["보류"] = counts.get("보류", 0) + 1

        else:  # 다시
            setuk = draft_setuk(hb)
            if already_ingested(m, setuk, reason):
                continue
            m.setdefault("rejects", []).append({"at": t, "reason": reason, "setuk": setuk})
            m["approved"], m["approved_at"] = False, None
            m["held"], m["held_at"] = False, None
            m["done"] = True                 # 다시 쓰라는 요청이므로 대상으로 남긴다
            m["material_at"] = t             # ← 이 줄이 '다시 써야 할 사람'으로 잡히게 한다
            m["updated"] = t
            changes.append((mp, m)); counts["다시"] = counts.get("다시", 0) + 1
            if not reason:
                # 까닭 없는 반려 = "같은 조건에서 한 번 더 돌려라". 프롬프트엔 아무것도 안 실린다.
                print(f"  · {hb} {name} — 까닭 없이 다시 쓰기(같은 조건에서 한 번 더 돌립니다)")

    print("\n반영할 것: " + (", ".join(f"{k} {v}명" for k, v in sorted(counts.items())) or "없음"))
    if not a.write:
        print("\n미리보기입니다. 실제로 반영하려면 --write 를 붙이세요.")
        return
    for mp, m in changes:
        with open(mp, "w", encoding="utf-8") as f:
            json.dump(m, f, ensure_ascii=False, indent=1)
    print(f"✅ {len(changes)}명 반영했습니다.")
    if counts.get("다시"):
        print("   다음: python3 tool/render_now.py --opus   (다시 쓸 사람만 잡힙니다)")


if __name__ == "__main__":
    main()
