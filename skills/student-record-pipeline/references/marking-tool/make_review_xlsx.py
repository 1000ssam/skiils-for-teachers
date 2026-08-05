#!/usr/bin/env python3
"""make_review_xlsx.py — 선생님이 보실 검수용 엑셀. 이 파이프라인이 내놓는 결과물은 이것 하나다.

세특 바로 옆에 **선생님이 직접 표시하신 대목**을 나란히 붙인다. 기계가 그 대목을
학생 글에서 찾아 맞춰보는 방식도 있지만, 그건 글자가 같은지만 볼 뿐이라 정작 위험한
것(뜻이 거꾸로 쓰인 경우 등)을 못 잡는다. 그래서 선생님이 눈으로 보시게 한다.

  python3 tool/make_review_xlsx.py [-o out/검수.xlsx]

⚠️ 이 파일은 나이스에 그대로 올리는 양식이 아니다(칸 구성이 다르다).
   내용을 확정하신 뒤 세특 칸만 나이스 양식으로 옮긴다.
🔒 학생이 쓴 글과 실명이 들어 있다. 깃에 올리지 말 것.
"""
import argparse
import glob
import json
import os
import re
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from gate import as_chars, check_text, load_rules, neis_bytes  # noqa: E402

# 책 인용 '제목(지은이)' — 책 정보는 학생도 기계도 자주 틀리게 적어 선생님 확인이 필요하다
BOOK = re.compile(r"'([^']{2,40})\(([^)]{2,20})\)'")

# 칸 너비 — '확인할 것' 이 가장 길어서 행 높이를 혼자 결정한다. 그래서 제일 넓게 준다.
COLS = [
    ("학번", 9), ("이름", 10), ("판정", 9), ("왜 다시 쓰나요", 30), ("세특", 58),
    ("선생님이 표시한 대목", 25), ("선생님이 정하신 것", 25), ("확인할 것", 60),
]
# '판정'·'왜 다시 쓰나요' 는 선생님이 채우는 칸이다(나머지는 읽는 칸).
# 학번·이름과 함께 틀고정 안에 둬서 오른쪽 칸을 읽는 동안에도 계속 보이게 한다.
# 채우신 뒤  python3 tool/ingest_review_xlsx.py  로 되먹이면 그 사유가 다음 초안에 실린다.
JUDGE_COLS = (3, 4)
HDR_FILL = PatternFill("solid", fgColor="2F5597")
JUDGE_FILL = PatternFill("solid", fgColor="EAF1FB")   # 선생님이 채우시는 칸
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
FAIL_FILL = PatternFill("solid", fgColor="FCE4E4")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load(hakbun_dir, drafts_dir):
    """학번 → (mark, draft).

    🔑 **초안이 없는 학생은 싣지 않는다.** 이 표는 '나온 문장을 판정하는 자리'다.
    아직 마킹 중이신 학생을 올리면 세특 칸이 비고 '고쳐야 할 것 — 초안 없음'이 붙어
    **붉은 줄**로 뜬다. 아직 안 하신 일이 결함처럼 보이는 것이라 헌법 8에 걸린다.
    (누가 남았는지는 render_now.py 가 상태표로 보여 준다 — 그게 그 일을 하는 자리다.)
    """
    rows = {}
    for fp in sorted(glob.glob(os.path.join(hakbun_dir, "*.json"))):
        hb = os.path.basename(fp)[:-5]
        if hb.startswith("_"):
            continue  # _context.json 등 메타파일
        rows.setdefault(hb, {})["mark"] = json.load(open(fp, encoding="utf-8"))
    for fp in sorted(glob.glob(os.path.join(drafts_dir, "*.json"))):
        hb = os.path.basename(fp)[:-5]
        rows.setdefault(hb, {})["draft"] = json.load(open(fp, encoding="utf-8"))
    rows = {hb: r for hb, r in rows.items() if r.get("draft")}
    return dict(sorted(rows.items()))


def prior_judgement(mark):
    """이미 정해 두신 판정을 되읽는다 — (판정, 까닭).

    🔑 이 표를 다시 만들 때 **선생님이 찍어 두신 확정·보류가 지워지면 안 된다.**
    '확정 → 남은 사람 렌더 → 표 다시 만들기' 는 정상 경로인데, 그때마다 판정이
    빈칸으로 돌아가면 같은 일을 두 번 하시게 된다(실측 2026-08-03).
    """
    if not mark:
        return "", ""
    if mark.get("approved"):
        return "통과", ""
    if mark.get("held"):
        return "보류", ""
    return "", ""


def fmt_spans(mark):
    """선생님이 표시하신 대목을 읽기 좋게. 표시가 없으면 그 사실만 적는다."""
    if not mark:
        return "(표시 기록이 없습니다)"
    hl = mark.get("highlights") or []
    if not hl:
        return "(표시 안 하심 — AI가 학생 글을 읽고 직접 골랐습니다)"
    out = []
    for i, h in enumerate(hl, 1):
        sec = h.get("section") or "?"
        out.append(f"{i}. [{sec}칸] {(h.get('text') or '').strip()}")
        if (why := (h.get("why") or "").strip()):
            out.append(f"    └ 선생님 메모: {why}")
    return "\n".join(out)


def fmt_teacher(mark):
    """선생님이 정하신 것 — 세특이 이걸 제대로 담았는지 옆에서 바로 견주시라고 붙인다."""
    if not mark:
        return ""
    out = [f"도달 수준: {mark.get('level') or '(안 정하심)'}"]
    if (c := (mark.get("competency") or "").strip()):
        out.append(f"보신 역량: {c}")
    if (e := (mark.get("extra") or "").strip()):
        out.append(f"당부하신 것: {e}")
    if rj := [r for r in (mark.get("rejects") or []) if r]:
        out.append(f"다시 쓰라고 하신 적 {len(rj)}번: "
                   + " / ".join(str(r)[:60] for r in rj[:3]))
    return "\n".join(out)


def fmt_notice(draft, hard, soft, nb, byte_max):
    """확인할 것 — 자동 검사 결과와, AI가 남긴 말."""
    out = [f"길이 {nb}바이트 (한글 약 {as_chars(nb)}자) / "
           f"최대 {byte_max}바이트 (약 {as_chars(byte_max)}자)"]
    out.append("고쳐야 할 것 — " + "; ".join(f"[{t}] {m}" for t, m in hard)
               if hard else "자동 검사 통과")
    for t, m in soft:
        out.append(f"⚠️ [{t}] {m}")
    if draft:
        if books := BOOK.findall(draft.get("setuk") or ""):
            out.append("📚 실제 책이 맞는지 확인해 주세요: "
                       + ", ".join(f"{t}({a})" for t, a in books))
        if (n := (draft.get("notes") or "").strip()):
            out.append(f"· AI가 이렇게 쓴 이유: {n}")
        if (u := (draft.get("unmet") or "").strip()):
            out.append(f"· 못 쓴 것과 그 까닭: {u}")
    return "\n".join(out)


def main():
    ap = argparse.ArgumentParser(description="선생님이 보실 검수용 엑셀 만들기")
    ap.add_argument("--marks", default="out/marks", help="선생님 표시 기록이 든 폴더")
    ap.add_argument("--drafts", default="out/drafts_v3b", help="세특 초안이 든 폴더")
    ap.add_argument("--rules", default="draft-rules.json", help="길이·못 쓰는 말 규칙")
    ap.add_argument("-o", "--out", default="out/검수.xlsx", help="만들어질 엑셀 파일")
    args = ap.parse_args()

    rules = load_rules(args.rules)
    byte_max = rules.get("byte_max", 1500)
    data = load(args.marks, args.drafts)
    if not data:
        sys.exit("검사 대상이 없다.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "세특 검수"

    for i, (title, width) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=i, value=title)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 24
    # 본문 행 높이는 정하지 않는다 — 엑셀이 글자량에 맞춰 스스로 잡게 둔다.
    # 계산해서 박아 본 적 있는데(글자 수 ÷ 칸 너비), 가장 긴 칸에 맞춰지는 건 똑같으면서
    # 짧은 학생은 절반이 빈칸이 됐다. 실물 비교 끝에 자동이 낫다고 판정(2026-08-02).
    ws.freeze_panes = "E2"      # 학번·이름·판정·왜 는 항상 보이게

    n_fail = n_deleg = 0
    for r, (hb, rec) in enumerate(data.items(), start=2):
        mark, draft = rec.get("mark"), rec.get("draft")
        setuk = (draft or {}).get("setuk") or ""
        hard, soft = check_text(setuk, rules) if setuk else ([("빈칸", "초안 없음")], [])
        nb = neis_bytes(setuk)
        deleg = bool(mark) and not (mark.get("highlights") or [])
        n_fail += bool(hard)
        n_deleg += deleg

        j_val, j_why = prior_judgement(mark)
        vals = [hb, (mark or {}).get("name", ""), j_val, j_why, setuk,
                fmt_spans(mark), fmt_teacher(mark),
                fmt_notice(draft, hard, soft, nb, byte_max)]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.alignment = Alignment(wrap_text=True, vertical="top",
                                    horizontal="center" if i <= 3 else "left")
            c.border = BORDER
            if hard:
                c.fill = FAIL_FILL
            elif deleg:
                c.fill = WARN_FILL
        for i in JUDGE_COLS:                    # 선생님이 채우실 칸은 잠그지 않고 눈에 띄게 둔다
            ws.cell(row=r, column=i).fill = JUDGE_FILL

    # '판정' 칸은 골라 넣게 한다 — 손으로 적으면 표기가 갈려(다시/재작성/X/O) 되먹일 때 못 읽는다.
    last = len(data) + 1
    if last >= 2:
        dv = DataValidation(type="list", formula1='"통과,재작성,보류"', allow_blank=True,
                            showDropDown=False)
        dv.error = "통과 · 재작성 · 보류 중에서 고르세요. 비워 두시면 아직 안 정하신 것으로 봅니다."
        dv.prompt = "이 세특을 어떻게 할까요? 비워 두면 '아직 안 정함'입니다."
        ws.add_data_validation(dv)
        dv.add(f"C2:C{last}")

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{last}"
    os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)

    # 선생님이 검수 엑셀을 열어 둔 채 마킹하시는 건 자연스러운 일이다. 그런데 엑셀이 파일을
    # 잠그기 때문에 그냥 저장하면 터진다 — 지켜보는 창(watch.sh)이 이걸 자동으로 부르므로
    # 그대로 두면 마킹할 때마다 오류가 난다. 그래서 다른 이름으로 흘려보낸다.
    out = args.out
    try:
        wb.save(out)
    except PermissionError:
        alt = f"{os.path.splitext(out)[0]}_새로고침.xlsx"
        try:
            wb.save(alt)
        except PermissionError:
            sys.exit(f"❌ {out} 과 {alt} 둘 다 열려 있어 저장하지 못했습니다.\n"
                     f"   엑셀에서 닫으신 뒤 다시 실행해 주세요.")
        print(f"⚠️ {out} 이 엑셀에서 열려 있어 덮어쓰지 못했습니다.")
        print(f"   대신 {alt} 로 저장했습니다. 엑셀을 닫으시면 다음부터 원래 이름으로 저장됩니다.")
        out = alt

    print(f"{len(data)}명 → {out}")
    print(f"  붉은 줄 = 고칠 것이 있는 학생: {n_fail}명")
    print(f"  노란 줄 = 선생님이 표시 안 하셔서 AI가 직접 고른 학생: {n_deleg}명")
    print("  ※ 나이스에 그대로 올리는 양식이 아닙니다. "
          "내용을 확정하신 뒤 세특 칸만 나이스 양식으로 옮깁니다.")


if __name__ == "__main__":
    main()
