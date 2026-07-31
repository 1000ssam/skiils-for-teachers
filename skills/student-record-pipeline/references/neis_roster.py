#!/usr/bin/env python3
"""NEIS 과목세특 내보내기 xlsx → 명렬표 정본 CSV (Step 1 Q7 ↔ Step 3 의 이음매).

왜 필요한가: 인테이크(Q7)는 "NEIS 과목세특 xlsx 를 주시면 최상"이라고 권한다 — 그 파일이
명렬표 정본이자 Step 7 write-back 타깃을 겸하기 때문이다. 그런데 하류 도구(map_pages.py·
build_students.py)는 CSV 명렬표를 전제한다. **그 사이를 잇는 단계가 없어서**, 교사가 권고대로
xlsx 만 주면 파이프라인이 거기서 끊기고 누군가 손으로 명렬표를 찾아 오게 된다(= 정본이 아닌
파일이 슬그머니 ground truth 가 되는 경로). 그 구멍을 막는다.

학번 = 학년(1) + 반(2) + 번호(2). 서식의 `반/번호`("4/1")와 `학년` 열에서 조립한다.
열은 위치가 아니라 **헤더명으로** 찾는다(서식 변동 내성).

사용:
    python3 neis_roster.py --targets "neis/*.xlsx" -o out/roster.csv
"""
import argparse, csv, glob, sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl 이 필요합니다:  pip install openpyxl")

H_CLSNUM, H_NAME, H_GRADE = "반/번호", "성명", "학년"


def col_index(ws, header):
    """헤더명 → 0-based 열 인덱스. 개행·공백 차이를 흡수한다(서식이 '학적변동\\n구분' 처럼 쓴다)."""
    row = [(c.value or "") for c in ws[1]]
    norm = ["".join(str(v).split()) for v in row]
    key = "".join(header.split())
    if key not in norm:
        raise KeyError(f"헤더 '{header}' 를 찾지 못했습니다. 실제 헤더: {row}")
    return norm.index(key)


def main():
    ap = argparse.ArgumentParser(description="NEIS 과목세특 xlsx 에서 명렬표 정본 CSV 를 뽑는다.")
    ap.add_argument("--targets", required=True, help='반별 xlsx 글롭. 예: "neis/*.xlsx"')
    ap.add_argument("-o", "--out", default="out/roster.csv")
    a = ap.parse_args()

    files = sorted(glob.glob(a.targets))
    if not files:
        sys.exit(f"대상 파일이 없습니다: {a.targets}")

    seen, rows = {}, []
    for f in files:
        ws = openpyxl.load_workbook(f, data_only=True).active
        ci_cn, ci_nm, ci_gr = col_index(ws, H_CLSNUM), col_index(ws, H_NAME), col_index(ws, H_GRADE)
        for r in range(2, ws.max_row + 1):
            cn = str(ws.cell(r, ci_cn + 1).value or "").strip()
            nm = str(ws.cell(r, ci_nm + 1).value or "").strip()
            gr = str(ws.cell(r, ci_gr + 1).value or "").strip()
            if not cn or not nm:
                continue
            cls_s, _, num_s = cn.partition("/")
            try:
                cls_i, num_i = int(cls_s), int(num_s)
            except ValueError:
                print(f"  ⚠ 건너뜀 — 반/번호 파싱 실패: {f} {cn!r} {nm}", file=sys.stderr)
                continue
            hakbun = f"{gr}{cls_i:02d}{num_i:02d}"
            key = (f"{gr}-{cls_i}", hakbun)
            if key in seen:
                # 같은 학생이 두 파일에 있으면(예: 한국사1·한국사2) 이름이 어긋나는지만 본다.
                if seen[key] != nm:
                    print(f"  ⚠ 같은 학번인데 이름이 다릅니다: {hakbun} {seen[key]} vs {nm} ({f})", file=sys.stderr)
                continue
            seen[key] = nm
            rows.append({"반": f"{gr}-{cls_i}", "학번": hakbun, "이름": nm, "_num": num_i})

    rows.sort(key=lambda r: (r["반"].split("-")[0], int(r["반"].split("-")[1]), r["_num"]))
    with open(a.out, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=["순서", "반", "학번", "이름"])
        w.writeheader()
        for i, r in enumerate(rows, 1):
            w.writerow({"순서": i, "반": r["반"], "학번": r["학번"], "이름": r["이름"]})

    by_cls = {}
    for r in rows:
        by_cls[r["반"]] = by_cls.get(r["반"], 0) + 1
    print(f"명렬표 정본 생성: {a.out} — {len(rows)}명 / {len(by_cls)}반 (원천 {len(files)}개 파일)")
    for c, n in sorted(by_cls.items()):
        print(f"  {c}: {n}명")
    print("\n이 CSV 가 Step3 매핑의 유일한 ground truth 다(헌법 1). 다른 명렬표를 섞지 말 것.")


if __name__ == "__main__":
    main()
